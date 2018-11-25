# -*- coding: utf-8 -*-
"""
Created on Sat Aug 26 15:06:28 2017

@author: Dave
"""

### Libraries
import openpyxl as op
import os
import pandas as pd
import pandas_datareader as pdr
import itertools
from datetime import datetime, timedelta

### Main Code

# Setup
WKdir = 'C:\\Users\\Dave\\Desktop\\'
os.chdir(WKdir)

# Open Aphid Data Feeds and pull in existing table
wb=op.load_workbook('Aphid_DF_test.xlsx')
sheets = wb.get_sheet_names()
sheet_SP500 = wb.get_sheet_by_name('qry_SP500')
last_row = sheet_SP500.max_row

# from https://openpyxl.readthedocs.io/en/default/pandas.html
data = sheet_SP500.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (itertools.islice(r, 1, None) for r in data)
df = pd.DataFrame(data, index=idx, columns=cols)

SP500_table = pd.DataFrame()

## Get current date
date_now = datetime.now()
date_last = idx[0] # Get most recent date from existing CSV file
date_delta = date_now-date_last
# convert object to the format we want
formatted_date = date_now.strftime('%Y-%m-%d')
print (formatted_date)
# SPY is an ETF that tracks the SP500 with data back to 1994.
# Multiply by 10 to get index value
# But volume data is different from volume of SP500
# Put this in a try/except loop
for i in range(0,date_delta.days):
    date_request = date_last + timedelta(i)
    print (date_request,'\n')
    try:
        SPY_yahoo = pdr.get_data_yahoo('SPY', date_request)[:1]
        print(SPY_yahoo)
    except:
        print ('Data not available for',date_request)
    

# Open Aphid Data Feeds and write latest SP500 data
wb=op.load_workbook('Aphid_DF_test.xlsx')
sheets = wb.get_sheet_names()
sheet_SP500 = wb.get_sheet_by_name('qry_SP500')
A1 = sheet_SP500['A1'].value
sheet_SP500['X1'] = 'Testing'
sheet_SP500['X1'].value
wb.save('Aphid_DF_test.xlsx')